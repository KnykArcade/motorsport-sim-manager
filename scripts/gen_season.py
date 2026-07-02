"""Generate TypeScript seed data for a historical F1 season from its
``F1_<year>_Season_Full_Update.xlsx`` workbook.

The workbooks are heterogeneous (different sheet names, header rows, and column
sets per year), so this is a tolerant, *header-name based* converter: it locates
each sheet by keyword, finds the header row by an anchor column, and reads values
by (normalized) column name with alias fallbacks. Fields the workbook does not
provide (e.g. driver salary in later years) are derived from ratings/age with
bounded formulas so the data stays internally consistent.

Like the 1995 generators, this is a build-time step. The emitted ``.ts`` files
are the runtime source of truth; the spreadsheets are never parsed at runtime.

Usage:
    python3 scripts/gen_season.py 1997 1998 1999 2000
    # or a single year:
    YEAR=1999 python3 scripts/gen_season.py
"""
import glob
import os
import re
import sys
import unicodedata

import openpyxl

ATTACH = os.environ.get("ATTACH", "/home/ubuntu/attachments")
OUT = os.environ.get("OUT", "/home/ubuntu/repos/motorsport-sim-manager/src/data")


# ----------------------------------------------------------------- primitives
def slug(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s


def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s).strip().lower()
    return re.sub(r"\s+", " ", s)


def num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace("\xa0", "").strip().rstrip("%")
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else None


def r1(x):
    return round(float(x) * 10) / 10


def ts_str(s):
    if s is None:
        s = ""
    s = str(s).replace("\xa0", " ").strip()
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def ts_bool(b):
    return "true" if b else "false"


WORD = {
    "very low": 1, "minimum": 1, "minimum-low": 2, "min": 1, "low": 3,
    "low-medium": 4, "medium-low": 4, "medium": 5, "med": 5, "balanced": 5,
    "medium-high": 6, "med-high": 6, "high": 7, "high-maximum": 8,
    "maximum": 9, "max": 9, "very high": 10, "elite": 9,
}


def wordnum(v, default=5):
    n = num(v)
    if n is not None:
        return r1(n)
    if v is None:
        return default
    return WORD.get(norm(v).replace(" ", "-"), WORD.get(norm(v), default))


# --------------------------------------------------------------- sheet access
class Sheet:
    """A worksheet wrapped with a normalized header->index map."""

    def __init__(self, ws, anchors):
        self.ws = ws
        self.rows = [list(r) for r in ws.iter_rows(values_only=True)]
        self.header_row = self._find_header(anchors)
        self.idx = {}
        if self.header_row is not None:
            for i, c in enumerate(self.rows[self.header_row]):
                if c is not None and str(c).strip():
                    self.idx.setdefault(norm(c), i)

    def _find_header(self, anchors):
        anchors = [norm(a) for a in anchors]
        for i, row in enumerate(self.rows[:10]):
            cells = [norm(c) for c in row if c is not None and str(c).strip()]
            if any(a in cells for a in anchors):
                return i
        return None

    def col(self, *aliases):
        for a in aliases:
            na = norm(a)
            if na in self.idx:
                return self.idx[na]
        # substring fallback
        for a in aliases:
            na = norm(a)
            for k, v in self.idx.items():
                if na == k or na in k.split():
                    return v
        return None

    def data_rows(self):
        if self.header_row is None:
            return []
        out = []
        for row in self.rows[self.header_row + 1:]:
            if not row:
                continue
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            out.append(row)
        return out


def get(row, sheet, *aliases):
    ci = sheet.col(*aliases)
    if ci is None or ci >= len(row):
        return None
    return row[ci]


def find_workbook(year, series):
    pat = f"{series}_{year}_Season_Full_Update"
    series_pat = f"{series}_"
    best, best_mtime = None, -1
    patterns = [
        os.path.join(ATTACH, "*", "*.xlsx"),
        os.path.join(ATTACH, "*.xlsx"),
    ]
    for pattern in patterns:
        for f in glob.glob(pattern):
            base = os.path.basename(f)
            if pat in base:
                mt = os.path.getmtime(f)
                if mt > best_mtime:
                    best, best_mtime = f, mt
            elif series_pat in base and "Season_Full_Update" in base:
                years_in_name = re.findall(r'(\d{4})', base)
                if str(year) in base:
                    mt = os.path.getmtime(f)
                    if mt > best_mtime:
                        best, best_mtime = f, mt
                elif len(years_in_name) >= 2:
                    lo, hi = int(years_in_name[0]), int(years_in_name[1])
                    if lo <= year <= hi:
                        mt = os.path.getmtime(f)
                        if mt > best_mtime:
                            best, best_mtime = f, mt
    return best


def pick_sheet(wb, *keyword_groups, year=None):
    """Return the first sheet name whose lowercased name matches all keywords in
    any of the keyword_groups (groups tried in order). If *year* is given, only
    sheets whose name contains that year string are considered (for multi-year
    workbooks). If no sheet matches with the year filter, falls back to
    searching without it (for single-year workbooks whose tabs lack the year)."""
    for group in keyword_groups:
        kws, excl = group[0], group[1] if len(group) > 1 else []
        for sn in wb.sheetnames:
            low = sn.lower()
            if year and str(year) not in low:
                continue
            if all(k in low for k in kws) and not any(e in low for e in excl):
                return sn
    # Fallback: retry without year filter if year-specific search found nothing
    if year:
        for group in keyword_groups:
            kws, excl = group[0], group[1] if len(group) > 1 else []
            for sn in wb.sheetnames:
                low = sn.lower()
                if all(k in low for k in kws) and not any(e in low for e in excl):
                    return sn
    return None


# ------------------------------------------------------------------- skills
SKILL_KEYS = [
    "cornering", "braking", "straights", "tractionAcceleration",
    "elevationBlindCorners", "technical", "overtakingRacecraft",
    "surfaceGripBumpiness", "riskManagement", "enduranceConsistency",
]
SKILL_ALIASES = {
    "cornering": ["corners", "cornering"],
    "braking": ["braking"],
    "straights": ["straights"],
    "tractionAcceleration": ["traction acceleration", "traction"],
    "elevationBlindCorners": ["elevation blind corner", "elevation blind corners", "elevation blind"],
    "technical": ["technical"],
    "overtakingRacecraft": ["overtaking racecraft", "overtaking", "racecraft"],
    "surfaceGripBumpiness": ["surface grip bumpiness", "surface bumpiness", "surface grip"],
    "riskManagement": ["risk management", "risk wall proximity", "risk wall", "crash avoidance"],
    "enduranceConsistency": ["endurance consistency", "endurance", "consistency"],
}


def read_skills(row, sheet, fallback):
    vals = {}
    present = []
    for k in SKILL_KEYS:
        v = num(get(row, sheet, *SKILL_ALIASES[k]))
        vals[k] = v
        if v is not None:
            present.append(v)
    base = round(sum(present) / len(present), 1) if present else fallback
    for k in SKILL_KEYS:
        if vals[k] is None:
            vals[k] = base
    return vals


def skills_ts(vals):
    return "{ " + ", ".join(f"{k}: {r1(vals[k])}" for k in SKILL_KEYS) + " }"


# --------------------------------------------------- team display metadata
TEAM_META = {
    "williams": ("WIL", "United Kingdom", "#1f6fd6"),
    "ferrari": ("FER", "Italy", "#d4001a"),
    "mclaren": ("MCL", "United Kingdom", "#ff8000"),
    "benetton": ("BEN", "United Kingdom", "#1fa05a"),
    "jordan": ("JOR", "Ireland", "#e6b800"),
    "sauber": ("SAU", "Switzerland", "#00a3d6"),
    "arrows": ("ARR", "United Kingdom", "#e25a2a"),
    "footwork": ("FTW", "United Kingdom", "#cc2a5e"),
    "minardi": ("MIN", "Italy", "#7a7f8a"),
    "prost": ("PRO", "France", "#2f6be0"),
    "ligier": ("LIG", "France", "#2f8fe0"),
    "stewart": ("STW", "United Kingdom", "#d6ffff"),
    "tyrrell": ("TYR", "United Kingdom", "#1f5fd0"),
    "lola": ("LOL", "United Kingdom", "#c0202a"),
    "bar": ("BAR", "United Kingdom", "#cf2030"),
    "jaguar": ("JAG", "United Kingdom", "#0b5d2e"),
    "forti": ("FOR", "Italy", "#d6c200"),
    "pacific": ("PAC", "United Kingdom", "#2aa198"),
    "larrousse": ("LAR", "France", "#1f3a8a"),
    "lotus": ("LOT", "United Kingdom", "#0b6b3a"),
    "simtek": ("SIM", "United Kingdom", "#5a5f6a"),
    # modern (2026) constructors, keyed by first slug token
    "mercedes": ("MER", "Germany", "#00a19c"),
    "red": ("RBR", "Austria", "#1e40af"),
    "alpine": ("ALP", "France", "#0090d0"),
    "racing": ("RB", "Italy", "#2b4cc0"),
    "haas": ("HAA", "United States", "#9aa0a6"),
    "audi": ("AUD", "Germany", "#bb0a30"),
    "aston": ("AMR", "United Kingdom", "#00665e"),
    "cadillac": ("CAD", "United States", "#c8a24a"),
    # IndyCar 2026 teams, matched by a distinctive name token
    "ganassi": ("CGR", "United States", "#e2231a"),
    "penske": ("PEN", "United States", "#d4a017"),
    "andretti": ("AND", "United States", "#1f6fd0"),
    "arrow": ("AMS", "United States", "#ff8000"),
    "meyer": ("MSR", "United States", "#1aa6a6"),
    "rahal": ("RLL", "United States", "#d11f3a"),
    "carpenter": ("ECR", "United States", "#0a3a8a"),
    "foyt": ("AJF", "United States", "#b08d2a"),
    "coyne": ("DCR", "United States", "#2a8a3a"),
    "juncos": ("JHR", "Argentina", "#5a2ad6"),
    "prema": ("PRE", "Italy", "#d6202a"),
}
PALETTE = ["#8a7ad6", "#d68a3a", "#3ad6c2", "#d63a8a", "#6ad63a", "#3a6ad6"]


def team_base(name):
    return slug(name).split("-")[0]


# Some workbooks name the same constructor differently in the car sheet vs the
# driver sheet within one season (e.g. "RBR Renault" vs "Red Bull Renault",
# "STR Ferrari" vs "Toro Rosso Ferrari"). Canonicalize so both resolve to the
# same team when seating drivers.
CANON_ALIASES = {
    "red bull": "redbull",
    "rbr": "redbull",
    "toro rosso": "tororosso",
    "str": "tororosso",
    "force india": "forceindia",
}


def canon_team(name):
    s = norm(name)
    for k, v in CANON_ALIASES.items():
        if k in s:
            return v
    return slug(name).split("-")[0]


def team_meta(name):
    s = slug(name)
    base = s.split("-")[0]
    if base in TEAM_META:
        return TEAM_META[base]
    tokens = s.split("-")
    for k, v in TEAM_META.items():
        if k in tokens:
            return v
    short = re.sub(r"[^A-Z]", "", name.upper())[:3] or name[:3].upper()
    color = PALETTE[hash(base) % len(PALETTE)]
    return (short, "", color)


# --------------------------------------------------- financial derivations
SALARY_TIER = {
    "elite": 10.0, "top": 8.0, "top tier": 8.5, "high": 7.0,
    "upper midfield": 5.0, "upper-midfield": 5.0, "midfield": 3.5,
    "lower midfield": 2.5, "lower-midfield": 2.5, "low": 2.0,
    "backmarker": 1.5, "rookie": 1.0, "minimum": 0.6, "very high": 9.0,
    "medium": 3.5, "medium-high": 6.0, "medium-low": 2.0,
}
BUDGET_TIER = {
    "elite": 95, "top": 85, "top tier": 88, "works": 80, "high": 75,
    "upper midfield": 62, "upper-midfield": 62, "midfield": 45,
    "lower midfield": 32, "lower-midfield": 32, "privateer": 30,
    "backmarker": 22, "low": 26, "minimum": 18,
}


def tier_value(v, table, scale=1.0):
    n = num(v)
    if n is not None:
        return n * scale
    if v is None:
        return None
    key = norm(v)
    if key in table:
        return table[key]
    for k, val in table.items():
        if k in key:
            return val
    return None


def to_millions(n):
    # Workbooks mix $M (e.g. 4.5) and absolute dollars (e.g. 4500000).
    return n / 1_000_000 if n is not None and n > 1000 else n


def derive_salary(overall, tier_raw):
    n = to_millions(num(tier_raw))
    if n is not None:
        return r1(n)
    t = tier_value(tier_raw, SALARY_TIER)
    if t is not None:
        return r1(t)
    return r1(max(0.4, (overall - 5) * 1.5 + 2))


def derive_sponsor(overall, raw):
    n = to_millions(num(raw))
    if n is not None:
        return r1(n)
    t = tier_value(raw, SALARY_TIER)
    if t is not None:
        return r1(max(0.3, t * 0.6))
    return r1(max(0.3, overall * 0.6))


def derive_buyout(overall, raw):
    n = to_millions(num(raw))
    if n is not None:
        return r1(n)
    return r1(max(0.0, (overall - 5.5) * 2.5))


def derive_devrate(age, raw):
    n = num(raw)
    if n is not None:
        return r1(n)
    if age is None:
        return 0.5
    return r1(min(1.0, max(0.2, (30 - age) * 0.05 + 0.3)))


def derive_readiness(overall, context, raw):
    n = num(raw)
    if n is not None:
        return int(round(min(100, max(0, n if n > 10 else n * 10))))
    base = overall * 10
    ctx = norm(context or "")
    if any(w in ctx for w in ["f3", "f3000", "feeder", "test", "junior", "karting", "gt"]):
        base -= 18
    return int(round(min(100, max(0, base))))


def derive_difficulty(overall, raw):
    if raw is not None:
        w = norm(raw)
        if w in ("low", "unwilling", "very hard", "hard"):
            return "Hard"
        if w in ("high", "very high", "willing", "easy", "eager"):
            return "Easy"
    if overall >= 8.5:
        return "Very Hard"
    if overall >= 7:
        return "Hard"
    if overall >= 5.5:
        return "Medium"
    return "Easy"


def points_for(y):
    """Historical F1 points system in effect for season year ``y``."""
    if y <= 1990:
        return "pts-1990"   # 9-6-4-3-2-1 (top 6)
    if y <= 2002:
        return "pts-1995"   # 10-6-4-3-2-1 (top 6)
    if y <= 2009:
        return "pts-2003"   # 10-8-6-5-4-3-2-1 (top 8)
    return "pts-modern"     # 25-18-15-... (top 10)


def yesno(v, default=False):
    if v is None:
        return default
    return norm(v) in ("yes", "true", "y", "eligible", "available")


# ------------------------------------------------------------------- emit
HEADER_TMPL = (
    "// AUTO-GENERATED from {src} by scripts/gen_season.py.\n"
    "// Edit ratings here directly; this file is the runtime source of truth.\n\n"
)


def gen(year, series="F1"):
    path = find_workbook(year, series)
    if not path:
        print(f"!! no workbook for {series} {year}")
        return
    src = os.path.basename(path)
    header = HEADER_TMPL.format(src=src)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    y = year
    tag = str(y) if series == "F1" else f"{y}{series}"

    cal_sn = pick_sheet(wb, (["calendar"],), (["track", "rating"],),
                        (["ratings"], ["key", "driver", "car"]), year=y)
    car_sn = pick_sheet(wb, (["car", "performance"],), (["team", "car"],),
                        (["carperformance"],), (["car"],), year=y)
    drv_sn = pick_sheet(wb, (["driver", "rating"],), (["drivers"],),
                        (["current"], ["market", "youth"]),
                        (["driver"], ["market"]), year=y)
    mkt_sn = pick_sheet(wb, (["driver", "market"],), (["market"],), year=y)
    yth_sn = pick_sheet(wb, (["youth"],), (["prospect"],), year=y)
    pts_sn = pick_sheet(wb, (["points"],), year=y)

    cal = Sheet(wb[cal_sn], ["round", "number"]) if cal_sn else None
    car = Sheet(wb[car_sn], ["team"]) if car_sn else None
    drv = Sheet(wb[drv_sn], ["driver"]) if drv_sn else None
    mkt = Sheet(wb[mkt_sn], ["driver", "pool"]) if mkt_sn else None
    yth = Sheet(wb[yth_sn], ["prospect", "name", "driver"]) if yth_sn else None

    # ---- tracks + calendar ----
    tracks = []
    for row in cal.data_rows():
        name = get(row, cal, "track")
        if not name:
            continue
        rnd = num(get(row, cal, "round", "number"))
        if rnd is None:
            continue
        attrs = {
            "corners": num(get(row, cal, *SKILL_ALIASES["cornering"])) or 5,
            "braking": num(get(row, cal, "braking")) or 5,
            "straights": num(get(row, cal, "straights")) or 5,
            "tractionAcceleration": num(get(row, cal, *SKILL_ALIASES["tractionAcceleration"])) or 5,
            "elevationBlindCorners": num(get(row, cal, *SKILL_ALIASES["elevationBlindCorners"])) or 5,
            "technical": num(get(row, cal, "technical")) or 5,
            "overtakingRacecraft": num(get(row, cal, *SKILL_ALIASES["overtakingRacecraft"])) or 5,
            "surfaceGripBumpiness": num(get(row, cal, *SKILL_ALIASES["surfaceGripBumpiness"])) or 5,
            "riskWallProximity": num(get(row, cal, "risk wall proximity", "risk wall", "risk management")) or 5,
            "enduranceConsistency": num(get(row, cal, *SKILL_ALIASES["enduranceConsistency"])) or 5,
        }
        tracks.append({
            "id": f"{slug(name)}-{tag}",
            "name": str(name).strip(),
            "gpName": str(get(row, cal, "gp name", "grand prix", "race") or name).strip(),
            "country": (str(get(row, cal, "country", "location") or "").strip()),
            "round": int(rnd),
            "km": num(get(row, cal, "track length km", "km", "track km")),
            "laps": num(get(row, cal, "race laps", "laps")),
            "archetype": str(get(row, cal, "track archetype") or "Balanced").strip(),
            "attrs": attrs,
            "setup": {
                "primarySetupProfile": str(get(row, cal, "primary setup profile") or "Balanced").strip(),
                "downforceLevel": str(get(row, cal, "downforce level", "downforce wing level", "recommended downforce") or "Medium").strip(),
                "topSpeedEmphasis": wordnum(get(row, cal, "top speed emphasis", "engine top speed emphasis")),
                "mechanicalGripEmphasis": wordnum(get(row, cal, "mechanical grip emphasis", "mechanical traction emphasis")),
                "brakeDemand": wordnum(get(row, cal, "brake demand")),
                "reliabilityRiskFocus": wordnum(get(row, cal, "reliability risk focus")),
                "strategyNotes": str(get(row, cal, "strategy notes", "strategy ops notes", "strategy setup notes") or "").strip(),
                "aeroDemand": num(get(row, cal, "aero demand")) or 5,
                "powerDemand": num(get(row, cal, "power demand")) or 5,
                "mechanicalDemand": num(get(row, cal, "mechanical demand")) or 5,
                "riskDemand": num(get(row, cal, "risk demand", "yellow caution risk")) or 5,
            },
            "notes": str(get(row, cal, "rating notes", "context", f"{y} context", "notes", "source notes") or "").strip(),
        })
    tracks.sort(key=lambda t: t["round"])

    # ---- teams + cars ----
    teams = []
    for row in car.data_rows():
        name = get(row, car, "team")
        if not name:
            continue
        name = str(name).strip()
        ratings = {
            "enginePower": scale10(get(row, car, "engine power")),
            "aeroEfficiency": scale10(get(row, car, "aero efficiency")),
            "mechanicalGrip": scale10(get(row, car, "mechanical grip")),
            "reliability": scale10(get(row, car, "reliability")),
            "pitCrewOperations": scale10(get(row, car, "pit crew operations")),
        }
        teams.append({
            "name": name,
            "ratings": ratings,
            "score": sum(ratings.values()),
            "budgetTier": get(row, car, "budget tier", "baseline tier"),
            "reputation": num(get(row, car, "reputation")),
        })
    teams.sort(key=lambda t: -t["score"])
    for i, t in enumerate(teams):
        t["rank"] = i + 1

    def budget_for(t):
        b = tier_value(t["budgetTier"], BUDGET_TIER)
        if b is None:
            b = max(20, 90 - (t["rank"] - 1) * 6)
        return int(round(b)) * 1_000_000

    def difficulty(rank):
        return "Easy" if rank <= 3 else ("Medium" if rank <= 6 else ("Hard" if rank <= 9 else "Very Hard"))

    team_by_slug = {slug(t["name"]): t for t in teams}
    team_by_canon = {canon_team(t["name"]): t for t in teams}

    # ---- drivers ----
    drivers = []
    for row in drv.data_rows():
        name = get(row, drv, "driver")
        if not name:
            continue
        name = str(name).strip()
        team_raw = str(get(row, drv, "team") or "").strip()
        overall = scale10(get(row, drv, "overall rating", "overall"), 6.0)
        sk = read_skills(row, drv, overall)
        qual = scale10(get(row, drv, "qualifying"),
                       r1((overall + sk["cornering"] + sk["braking"]) / 3))
        pace = scale10(get(row, drv, "race pace"),
                       r1((overall + sk["enduranceConsistency"] + sk["tractionAcceleration"]) / 3))
        adaptability = r1((overall + sk["technical"] + sk["surfaceGripBumpiness"]) / 3)
        aggression = r1((sk["overtakingRacecraft"] + (11 - sk["riskManagement"])) / 2)
        composure = r1((sk["riskManagement"] + sk["enduranceConsistency"] + overall) / 3)
        drivers.append({
            "id": f"d-{y}-{slug(name)}",
            "name": name,
            "team_slug": slug(team_raw),
            "team_raw": team_raw,
            "age": num(get(row, drv, f"age in {y}", "age")),
            "nationality": str(get(row, drv, "nationality") or "").strip(),
            "skills": sk,
            "qualifying": qual, "racePace": pace, "adaptability": adaptability,
            "aggression": aggression, "composure": composure, "overall": overall,
        })

    # assign drivers to teams (best 2 by overall whose team matches)
    for t in teams:
        t["driver_ids"] = []
    unassigned = []
    for d in sorted(drivers, key=lambda x: -x["overall"]):
        tslug = d["team_slug"]
        team = team_by_slug.get(tslug) or team_by_canon.get(canon_team(d["team_raw"]))
        if team is None:
            for s, tm in team_by_slug.items():
                if s.split("-")[0] == tslug.split("-")[0] and tslug:
                    team = tm
                    break
        if team is not None and len(team["driver_ids"]) < 2:
            team["driver_ids"].append(d["id"])
            d["team_raw"] = team["name"]
        else:
            unassigned.append(d["name"])
    grid_ids = {i for t in teams for i in t["driver_ids"]}
    drivers = [d for d in drivers if d["id"] in grid_ids]
    drivers.sort(key=lambda d: -d["overall"])
    for n, d in enumerate(drivers):
        d["number"] = n + 1

    # ---- market ----
    market = []
    if mkt:
        for row in mkt.data_rows():
            name = get(row, mkt, "driver")
            if not name:
                continue
            name = str(name).strip()
            overall = scale10(get(row, mkt, "overall rating", "overall"), 6.0)
            sk = read_skills(row, mkt, overall)
            context = str(get(row, mkt, "current series status", "current series context", "current series", "current series role", "current status") or "").strip()
            pool = str(get(row, mkt, "market pool", "pool") or "Senior").strip()
            status = str(get(row, mkt, "market status") or pool).strip()
            role = str(get(row, mkt, "likely role", "suggested role", "primary role", "suggested use") or "").strip()
            potential = scale10(get(row, mkt, "potential rating"), r1(min(10, overall + 0.4)))
            market.append({
                "id": f"mkt-{y}-{slug(name)}",
                "name": name,
                "age": int(num(get(row, mkt, f"age in {y}", "age")) or 0),
                "nationality": str(get(row, mkt, "nationality") or "").strip(),
                "context": context, "pool": pool, "status": status, "role": role,
                "eligible": yesno(get(row, mkt, "immediate f1 eligible"), overall >= 6.5 and "senior" in pool.lower()),
                "skills": sk, "overall": overall, "potential": potential,
                "potentialDelta": r1(potential - overall),
                "devRate": derive_devrate(num(get(row, mkt, f"age in {y}", "age")), get(row, mkt, "development rate")),
                "readiness": derive_readiness(overall, context, get(row, mkt, "f1 readiness")),
                "salary": derive_salary(overall, get(row, mkt, "salary estimate", "estimated salary", "salary demand", "salary tier", "salary demand tier", "estimated cost", "salary")),
                "sponsor": derive_sponsor(overall, get(row, mkt, "sponsor value")),
                "buyout": derive_buyout(overall, get(row, mkt, "buyout approach cost", "buyout")),
                "negDiff": derive_difficulty(overall, get(row, mkt, "negotiation difficulty", "willingness to move")),
                "use": role,
                "notes": str(get(row, mkt, "game notes", "notes") or "").strip(),
            })

    # ---- youth ----
    youth = []
    if yth:
        for row in yth.data_rows():
            name = get(row, yth, "name", "prospect", "driver")
            if not name:
                continue
            name = str(name).strip()
            overall = scale10(get(row, yth, "overall youth rating", "overall rating", "potential rating"), 6.0)
            sk = read_skills(row, yth, overall)
            rawpace = num(get(row, yth, "raw pace"))
            if rawpace is not None:
                sk["cornering"] = sk["cornering"] if get(row, yth, "corners") else rawpace
                sk["straights"] = sk["straights"] if get(row, yth, "straights") else rawpace
            race = num(get(row, yth, "racecraft"))
            if race is not None and not get(row, yth, "overtaking"):
                sk["overtakingRacecraft"] = race
            age = num(get(row, yth, f"age in {y}", "age"))
            potential = scale10(get(row, yth, "potential rating"), r1(min(10, overall + 1)))
            youth.append({
                "id": f"yth-{y}-{slug(name)}",
                "name": name,
                "age": int(age or 0),
                "birthYear": int(num(get(row, yth, "birth year")) or (y - (age or 17))),
                "nationality": str(get(row, yth, "nationality") or "").strip(),
                "level": str(get(row, yth, "current level") or "Karting").strip(),
                "pool": str(get(row, yth, "market pool") or "Youth").strip(),
                "status": str(get(row, yth, "market status", "academy eligibility") or "Prospect").strip(),
                "eligible": yesno(get(row, yth, "academy eligible now", "academy eligibility"), True),
                "earliest": int(num(get(row, yth, "earliest full academy year")) or y),
                "skills": sk, "overall": overall, "potential": potential,
                "potentialDelta": r1(potential - overall),
                "devRate": derive_devrate(age, get(row, yth, "development rate")),
                "years": int(num(get(row, yth, "years until f1 ready", "years until f1 ready estimate", "years until indycar ready")) or max(1, 18 - int(age or 17))),
                # Youth are unproven: derive cheap, consistent $M costs from
                # potential rather than the inconsistent source columns (some
                # sheets store $M, others raw dollars). Mirrors the runtime
                # normalization in src/data/market/index.ts.
                "signing": r1(0.02 + (max(0.0, min(10.0, potential)) / 10) * 0.13),
                "academy": r1(0.01 + (max(0.0, min(10.0, potential)) / 10) * 0.09),
                "risk": str(get(row, yth, "risk level") or "Medium").strip(),
                "path": str(get(row, yth, "suggested path") or "Academy").strip(),
                "notes": str(get(row, yth, "game notes", "notes", "traits") or "").strip(),
            })

    wb.close()
    points_id = points_for(y)
    write_ts(y, src, header, tracks, teams, drivers, market, youth, budget_for,
             difficulty, series=series, points_id=points_id)
    print(f"OK {series} {y}: tracks={len(tracks)} teams={len(teams)} drivers={len(drivers)} market={len(market)} youth={len(youth)}")
    if unassigned:
        print(f"   (off-grid drivers not seated: {unassigned})")


def scale10(v, default=5.0):
    """Normalize a rating to a 0-10 scale (workbooks mix 0-10 and 0-100)."""
    n = num(v)
    if n is None:
        return default
    return r1(n / 10) if n > 10 else r1(n)


def avg(*xs):
    xs = [x for x in xs if x is not None]
    return r1(sum(xs) / len(xs)) if xs else 5.0


def indy_skills(row, sheet):
    """Map the IndyCar driver skill columns onto the F1 10-skill model.

    Oval aptitude drives top-speed/traction skills; road/street aptitude drives
    cornering/technical; the remaining columns map by closest analogue.
    """
    oval = num(get(row, sheet, "oval")) or 5
    road = num(get(row, sheet, "road street", "road")) or 5
    racecraft = num(get(row, sheet, "racecraft")) or 5
    crash = num(get(row, sheet, "crash avoidance")) or 5
    consistency = num(get(row, sheet, "consistency")) or 5
    return {
        "cornering": r1(road),
        "braking": avg(road, oval),
        "straights": r1(oval),
        "tractionAcceleration": r1(oval),
        "elevationBlindCorners": r1(road),
        "technical": r1(road),
        "overtakingRacecraft": r1(racecraft),
        "surfaceGripBumpiness": avg(oval, road),
        "riskManagement": r1(crash),
        "enduranceConsistency": r1(consistency),
    }


def gen_indycar(year=2026):
    series = "IndyCar"
    path = find_workbook(year, series)
    if not path:
        print(f"!! no workbook for {series} {year}")
        return
    src = os.path.basename(path)
    header = HEADER_TMPL.format(src=src)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    y, tag = year, f"{year}{series}"

    cal = Sheet(wb[pick_sheet(wb, (["calendar"],))], ["round"])
    setup_sn = pick_sheet(wb, (["setup", "profile"],), (["track", "setup"],))
    setup = Sheet(wb[setup_sn], ["round"]) if setup_sn else None
    car = Sheet(wb[pick_sheet(wb, (["team", "car"],), (["teams", "cars"],))], ["team"])
    drv = Sheet(wb[pick_sheet(wb, (["driver"], ["market", "youth"]), (["drivers"],))], ["driver"])
    mkt = Sheet(wb[pick_sheet(wb, (["driver", "market"],), (["market"],))], ["driver"])
    yth = Sheet(wb[pick_sheet(wb, (["youth"],), (["prospect"],))], ["prospect", "name"])

    # ---- setup profiles keyed by round ----
    setup_by_round = {}
    if setup:
        for row in setup.data_rows():
            rnd = num(get(row, setup, "round"))
            if rnd is not None:
                setup_by_round[int(rnd)] = row

    # ---- tracks + calendar (distances are in miles) ----
    MI_KM = 1.60934
    tracks = []
    for row in cal.data_rows():
        name = get(row, cal, "track")
        rnd = num(get(row, cal, "round"))
        if not name or rnd is None:
            continue
        rnd = int(rnd)
        attrs = {
            "corners": num(get(row, cal, "corners")) or 5,
            "braking": num(get(row, cal, "braking")) or 5,
            "straights": num(get(row, cal, "straights")) or 5,
            "tractionAcceleration": num(get(row, cal, "traction")) or 5,
            "elevationBlindCorners": num(get(row, cal, "elevation blind", "elevation_blind")) or 5,
            "technical": num(get(row, cal, "technical")) or 5,
            "overtakingRacecraft": num(get(row, cal, "overtaking")) or 5,
            "surfaceGripBumpiness": num(get(row, cal, "surface bumpiness", "surface_bumpiness")) or 5,
            "riskWallProximity": num(get(row, cal, "risk wall", "risk_wall")) or 5,
            "enduranceConsistency": num(get(row, cal, "endurance")) or 5,
        }
        miles = num(get(row, cal, "track length mi", "track_length_mi"))
        sp = setup_by_round.get(rnd)

        def sget(*a):
            return get(sp, setup, *a) if (sp and setup) else None

        ttype = str(get(row, cal, "track type", "track_type") or "").strip()
        tracks.append({
            "id": f"{slug(name)}-{tag}",
            "name": str(name).strip(),
            "gpName": str(get(row, cal, "race") or name).strip(),
            "country": str(get(row, cal, "location") or "United States").strip(),
            "round": rnd,
            "km": r1(miles * MI_KM) if miles else None,
            "laps": num(get(row, cal, "laps sim default", "laps_sim_default")),
            "archetype": str(get(row, cal, "track archetype", "track_archetype") or ttype or "Balanced").strip(),
            "attrs": attrs,
            "setup": {
                "primarySetupProfile": ttype or "Balanced",
                "downforceLevel": str(sget("recommended downforce", "recommended_downforce") or "Medium").strip(),
                "topSpeedEmphasis": wordnum(sget("top speed emphasis", "top_speed_emphasis")),
                "mechanicalGripEmphasis": wordnum(sget("mechanical grip emphasis", "mechanical_grip_emphasis")),
                "brakeDemand": wordnum(sget("brake demand", "brake_demand")),
                "reliabilityRiskFocus": wordnum(sget("tire degradation", "tire_degradation")),
                "strategyNotes": str(sget("recommended strategy", "recommended_strategy") or get(row, cal, "strategy notes", "strategy_notes") or "").strip(),
                "aeroDemand": num(get(row, cal, "aero demand", "aero_demand")) or 5,
                "powerDemand": num(get(row, cal, "power demand", "power_demand")) or 5,
                "mechanicalDemand": num(get(row, cal, "mechanical demand", "mechanical_demand")) or 5,
                "riskDemand": num(get(row, cal, "yellow caution risk", "yellow_caution_risk")) or 5,
            },
            "notes": str(get(row, cal, "strategy notes", "strategy_notes") or "").strip(),
        })
    tracks.sort(key=lambda t: t["round"])

    # ---- teams + cars ----
    teams = []
    for row in car.data_rows():
        name = get(row, car, "team")
        if not name:
            continue
        name = str(name).strip()
        ratings = {
            "enginePower": scale10(get(row, car, "engineering")),
            "aeroEfficiency": scale10(get(row, car, "road street setup", "road_street_setup")),
            "mechanicalGrip": scale10(get(row, car, "oval setup", "oval_setup")),
            "reliability": scale10(get(row, car, "reliability")),
            "pitCrewOperations": scale10(get(row, car, "pit crew", "pit_crew")),
        }
        teams.append({
            "name": name,
            "ratings": ratings,
            "score": num(get(row, car, "overall")) or sum(ratings.values()),
            "devBudget": num(get(row, car, "development budget", "development_budget")),
            "reputation": num(get(row, car, "sponsor strength", "sponsor_strength")),
        })
    teams.sort(key=lambda t: -t["score"])
    for i, t in enumerate(teams):
        t["rank"] = i + 1
    team_by_slug = {slug(t["name"]): t for t in teams}

    def budget_for(t):
        b = t["devBudget"]
        b = (b * 10) if b else max(20, 90 - (t["rank"] - 1) * 6)
        return int(round(b)) * 1_000_000

    def difficulty(rank):
        return "Easy" if rank <= 3 else ("Medium" if rank <= 6 else ("Hard" if rank <= 9 else "Very Hard"))

    # ---- drivers (full-time entrants only) ----
    drivers = []
    for row in drv.data_rows():
        name = get(row, drv, "driver")
        if not name:
            continue
        status = norm(get(row, drv, "status") or "")
        if status and "full" not in status:
            continue
        name = str(name).strip()
        team_raw = str(get(row, drv, "team") or "").strip()
        if not team_raw:
            continue
        overall = scale10(get(row, drv, "overall"), 6.0)
        sk = indy_skills(row, drv)
        oval = num(get(row, drv, "oval")) or 5
        road = num(get(row, drv, "road street", "road")) or 5
        drivers.append({
            "id": f"d-{tag}-{slug(name)}",
            "name": name,
            "team_slug": slug(team_raw),
            "team_raw": team_raw,
            "age": num(get(row, drv, "age 2026", "age_2026", "age")),
            "nationality": str(get(row, drv, "nationality") or "").strip(),
            "skills": sk,
            "qualifying": scale10(get(row, drv, "qualifying"), overall),
            "racePace": scale10(get(row, drv, "race pace", "race_pace"), overall),
            "adaptability": avg(oval, road),
            "aggression": num(get(row, drv, "passing defense", "passing_defense")) or 5,
            "composure": num(get(row, drv, "consistency")) or 5,
            "overall": overall,
        })

    # assign every full-time driver to its team (IndyCar runs 2-4 cars/team)
    for t in teams:
        t["driver_ids"] = []
    unassigned = []
    for d in sorted(drivers, key=lambda x: -x["overall"]):
        team = team_by_slug.get(d["team_slug"])
        if team is None:
            for s, tm in team_by_slug.items():
                if s.split("-")[0] == d["team_slug"].split("-")[0] and d["team_slug"]:
                    team = tm
                    break
        if team is not None:
            team["driver_ids"].append(d["id"])
            d["team_raw"] = team["name"]
        else:
            unassigned.append(d["name"])
    teams = [t for t in teams if t["driver_ids"]]
    for i, t in enumerate(teams):
        t["rank"] = i + 1
    grid_ids = {i for t in teams for i in t["driver_ids"]}
    drivers = [d for d in drivers if d["id"] in grid_ids]
    drivers.sort(key=lambda d: -d["overall"])
    for n, d in enumerate(drivers):
        d["number"] = n + 1

    # ---- market (no per-skill columns; skills derived from overall) ----
    market = []
    for row in mkt.data_rows():
        name = get(row, mkt, "driver")
        if not name:
            continue
        name = str(name).strip()
        overall = scale10(get(row, mkt, "overall"), 6.0)
        sk = read_skills(row, mkt, overall)
        context = str(get(row, mkt, "current status", "current_status") or "").strip()
        status = str(get(row, mkt, "market status", "market_status") or "Senior").strip()
        potential = scale10(get(row, mkt, "potential"), r1(min(10, overall + 0.4)))
        age = int(num(get(row, mkt, "age 2026", "age_2026", "age")) or 0)
        market.append({
            "id": f"mkt-{tag}-{slug(name)}",
            "name": name, "age": age,
            "nationality": str(get(row, mkt, "nationality") or "").strip(),
            "context": context, "pool": "Senior", "status": status, "role": "",
            "eligible": overall >= 6.5,
            "skills": sk, "overall": overall, "potential": potential,
            "potentialDelta": r1(potential - overall),
            "devRate": derive_devrate(age, None),
            "readiness": derive_readiness(overall, context, get(row, mkt, "oval readiness", "oval_readiness")),
            "salary": derive_salary(overall, get(row, mkt, "estimated cost", "estimated_cost")),
            "sponsor": derive_sponsor(overall, None),
            "buyout": derive_buyout(overall, get(row, mkt, "estimated cost", "estimated_cost")),
            "negDiff": derive_difficulty(overall, get(row, mkt, "willingness to move", "willingness_to_move")),
            "use": "", "notes": str(get(row, mkt, "notes") or "").strip(),
        })

    # ---- youth ----
    youth = []
    for row in yth.data_rows():
        name = get(row, yth, "name", "prospect")
        if not name:
            continue
        name = str(name).strip()
        oval = num(get(row, yth, "oval aptitude", "oval_aptitude")) or 5
        road = num(get(row, yth, "road street aptitude", "road_street_aptitude")) or 5
        rawpace = num(get(row, yth, "raw pace", "raw_pace")) or 5
        racecraft = num(get(row, yth, "racecraft")) or 5
        composure = num(get(row, yth, "composure")) or 5
        sk = {
            "cornering": r1(road), "braking": avg(road, oval), "straights": r1(oval),
            "tractionAcceleration": r1(oval), "elevationBlindCorners": r1(road),
            "technical": r1(road), "overtakingRacecraft": r1(racecraft),
            "surfaceGripBumpiness": avg(oval, road), "riskManagement": r1(composure),
            "enduranceConsistency": r1(composure),
        }
        overall = avg(oval, road, rawpace, racecraft)
        potential = scale10(get(row, yth, "potential"), r1(min(10, overall + 1)))
        age = num(get(row, yth, "age 2026", "age_2026", "age"))
        youth.append({
            "id": f"yth-{tag}-{slug(name)}",
            "name": name, "age": int(age or 0),
            "birthYear": int(y - (age or 17)),
            "nationality": str(get(row, yth, "nationality") or "").strip(),
            "level": str(get(row, yth, "current level", "current_level") or "Karting").strip(),
            "pool": "Youth", "status": "Prospect", "eligible": True, "earliest": y,
            "skills": sk, "overall": overall, "potential": potential,
            "potentialDelta": r1(potential - overall),
            "devRate": derive_devrate(age, get(row, yth, "development rate", "development_rate")),
            "years": int(num(get(row, yth, "years until indycar ready", "years_until_indycar_ready")) or max(1, 18 - int(age or 17))),
            # Cheap, potential-derived $M costs (see note in the F1 youth block).
            "signing": r1(0.02 + (max(0.0, min(10.0, potential)) / 10) * 0.13),
            "academy": r1(0.01 + (max(0.0, min(10.0, potential)) / 10) * 0.09),
            "risk": str(get(row, yth, "risk level", "risk_level") or "Medium").strip(),
            "path": "Academy",
            "notes": str(get(row, yth, "notes", "traits") or "").strip(),
        })

    wb.close()
    write_ts(y, src, header, tracks, teams, drivers, market, youth, budget_for,
             difficulty, series=series, points_id="pts-indycar-2026")
    print(f"OK {series} {y}: tracks={len(tracks)} teams={len(teams)} drivers={len(drivers)} market={len(market)} youth={len(youth)}")
    if unassigned:
        print(f"   (drivers not seated: {unassigned})")


SERIES_NAME = {"F1": "Formula 1 World Championship", "IndyCar": "IndyCar Series"}


def write_ts(y, src, header, tracks, teams, drivers, market, youth, budget_for,
             difficulty, series="F1", points_id="pts-1995"):
    tag = str(y) if series == "F1" else f"{y}{series}"

    def ensure(*parts):
        d = os.path.join(OUT, *parts)
        os.makedirs(os.path.dirname(d), exist_ok=True)
        return d

    # tracks
    with open(ensure("tracks", f"tracks{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { Track } from '../../types/gameTypes';\n\n")
        f.write(f"export const tracks{tag}: Track[] = [\n")
        for t in tracks:
            a, s = t["attrs"], t["setup"]
            f.write("  {\n")
            f.write(f"    id: {ts_str(t['id'])},\n    name: {ts_str(t['name'])},\n")
            f.write(f"    gpName: {ts_str(t['gpName'])},\n")
            if t["country"]:
                f.write(f"    country: {ts_str(t['country'])},\n")
            f.write(f"    archetype: {ts_str(t['archetype'])},\n")
            f.write("    attributes: {\n")
            for k, v in a.items():
                f.write(f"      {k}: {r1(v)},\n")
            f.write("    },\n    setupProfile: {\n")
            f.write(f"      primarySetupProfile: {ts_str(s['primarySetupProfile'])},\n")
            f.write(f"      downforceLevel: {ts_str(s['downforceLevel'])},\n")
            f.write(f"      topSpeedEmphasis: {r1(s['topSpeedEmphasis'])},\n")
            f.write(f"      mechanicalGripEmphasis: {r1(s['mechanicalGripEmphasis'])},\n")
            f.write(f"      brakeDemand: {r1(s['brakeDemand'])},\n")
            f.write(f"      reliabilityRiskFocus: {r1(s['reliabilityRiskFocus'])},\n")
            f.write(f"      strategyNotes: {ts_str(s['strategyNotes'])},\n")
            f.write(f"      aeroDemand: {r1(s['aeroDemand'])},\n")
            f.write(f"      powerDemand: {r1(s['powerDemand'])},\n")
            f.write(f"      mechanicalDemand: {r1(s['mechanicalDemand'])},\n")
            f.write(f"      riskDemand: {r1(s['riskDemand'])},\n")
            f.write("    },\n")
            f.write(f"    ratingNotes: {ts_str(t['notes'])},\n  }},\n")
        f.write("];\n")

    # season + calendar
    with open(ensure("seasons", f"season{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { Race, Season } from '../../types/gameTypes';\n")
        f.write(f"import {{ tracks{tag} }} from '../tracks/tracks{tag}';\n\n")
        f.write(f"export const calendar{tag}: Race[] = [\n")
        for t in tracks:
            dist = r1(t["km"] * t["laps"]) if (t["km"] and t["laps"]) else "undefined"
            f.write("  {\n")
            f.write(f"    id: {ts_str(f'r-{y}-' + str(t['round']))},\n")
            f.write(f"    round: {t['round']},\n")
            f.write(f"    gpName: {ts_str(t['gpName'])},\n")
            f.write(f"    trackId: {ts_str(t['id'])},\n")
            f.write(f"    trackName: {ts_str(t['name'])},\n")
            f.write(f"    laps: {int(t['laps']) if t['laps'] else 0},\n")
            f.write(f"    distanceKm: {dist},\n    completed: false,\n  }},\n")
        f.write("];\n\n")
        f.write(f"void tracks{tag};\n\n")
        sid = f"s-{y}-{series.lower()}"
        sname = f"{y} {SERIES_NAME.get(series, 'Championship')}"
        f.write(f"export const season{tag}: Season = {{\n")
        f.write(f"  id: {ts_str(sid)},\n  year: {y},\n  name: {ts_str(sname)},\n")
        f.write(f"  series: {ts_str(series)},\n")
        f.write(f"  calendar: calendar{tag},\n  pointsSystemId: {ts_str(points_id)},\n  regulationSetId: 'reg-1995',\n}};\n")

    # teams
    with open(ensure("teams", f"teams{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { Team } from '../../types/gameTypes';\n\n")
        f.write(f"export const teams{tag}: Team[] = [\n")
        for t in teams:
            short, country, color = team_meta(t["name"])
            rank = t["rank"]
            raw_rep = t["reputation"]
            if raw_rep:
                rep = int(round(raw_rep * 10)) if raw_rep <= 10 else int(round(raw_rep))
                rep = min(100, max(10, rep))
            else:
                rep = max(20, 100 - (rank - 1) * 7)
            ids = ", ".join(ts_str(i) for i in t["driver_ids"])
            f.write("  {\n")
            f.write(f"    id: {ts_str('t-' + slug(t['name']))},\n")
            f.write(f"    name: {ts_str(t['name'])},\n")
            f.write(f"    shortName: {ts_str(short)},\n")
            if country:
                f.write(f"    country: {ts_str(country)},\n")
            f.write(f"    carId: {ts_str('car-' + slug(t['name']) + f'-{y}')},\n")
            f.write(f"    driverIds: [{ids}],\n")
            f.write(f"    budget: {budget_for(t)},\n")
            f.write(f"    reputation: {rep},\n")
            cr = t.get("ratings") or {}
            pit = cr.get("pitCrewOperations")
            rel = cr.get("reliability")
            if pit is not None and rel is not None:
                ops = 0.55 * (rep / 10) + 0.45 * ((pit + rel) / 2)
            else:
                ops = rep / 10
            ops = max(1.0, min(10.0, round(ops * 10) / 10))
            ops_s = str(int(ops)) if ops == int(ops) else f"{ops:g}"
            f.write(f"    raceOperations: {ops_s},\n    morale: 65,\n")
            f.write(f"    expectedStanding: {rank},\n")
            f.write(f"    difficulty: {ts_str(difficulty(rank))},\n")
            f.write(f"    color: {ts_str(color)},\n  }},\n")
        f.write("];\n")

    # cars
    with open(ensure("cars", f"cars{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { Car } from '../../types/gameTypes';\n\n")
        f.write(f"export const cars{tag}: Car[] = [\n")
        for t in teams:
            c = t["ratings"]
            f.write("  {\n")
            f.write(f"    id: {ts_str('car-' + slug(t['name']) + f'-{y}')},\n")
            f.write(f"    teamId: {ts_str('t-' + slug(t['name']))},\n")
            f.write(f"    seasonYear: {y},\n    ratings: {{\n")
            for k in ["enginePower", "aeroEfficiency", "mechanicalGrip", "reliability", "pitCrewOperations"]:
                f.write(f"      {k}: {r1(c[k])},\n")
            f.write("    },\n    condition: 100,\n")
            f.write("    developmentLevel: { enginePower: 0, aeroEfficiency: 0, mechanicalGrip: 0, reliability: 0, pitCrewOperations: 0 },\n  },\n")
        f.write("];\n")

    # drivers
    with open(ensure("drivers", f"drivers{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { Driver } from '../../types/gameTypes';\n\n")
        f.write(f"export const drivers{tag}: Driver[] = [\n")
        for d in drivers:
            sk = d["skills"]
            f.write("  {\n")
            f.write(f"    id: {ts_str(d['id'])},\n    name: {ts_str(d['name'])},\n")
            f.write(f"    number: {d['number']},\n")
            if d["nationality"]:
                f.write(f"    nationality: {ts_str(d['nationality'])},\n")
            if d["age"]:
                f.write(f"    age: {int(d['age'])},\n")
            f.write(f"    teamId: {ts_str('t-' + slug(d['team_raw']))},\n")
            f.write("    ratings: {\n")
            f.write(f"      cornering: {r1(sk['cornering'])},\n")
            f.write(f"      braking: {r1(sk['braking'])},\n")
            f.write(f"      straights: {r1(sk['straights'])},\n")
            f.write(f"      tractionAcceleration: {r1(sk['tractionAcceleration'])},\n")
            f.write(f"      elevationBlindCorners: {r1(sk['elevationBlindCorners'])},\n")
            f.write(f"      technical: {r1(sk['technical'])},\n")
            f.write(f"      overtakingRacecraft: {r1(sk['overtakingRacecraft'])},\n")
            f.write(f"      surfaceGripBumpiness: {r1(sk['surfaceGripBumpiness'])},\n")
            f.write(f"      riskManagement: {r1(sk['riskManagement'])},\n")
            f.write(f"      enduranceConsistency: {r1(sk['enduranceConsistency'])},\n")
            f.write(f"      qualifying: {r1(d['qualifying'])},\n")
            f.write(f"      racePace: {r1(d['racePace'])},\n")
            f.write(f"      adaptability: {r1(d['adaptability'])},\n")
            f.write(f"      aggression: {r1(d['aggression'])},\n")
            f.write(f"      composure: {r1(d['composure'])},\n")
            f.write(f"      overall: {r1(d['overall'])},\n")
            f.write("    },\n    morale: 65,\n    confidence: 65,\n    traits: [],\n  },\n")
        f.write("];\n")

    # market
    with open(ensure("market", f"driverMarket{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { MarketDriver } from '../../types/marketTypes';\n\n")
        f.write(f"export const driverMarket{tag}: MarketDriver[] = [\n")
        for m in market:
            f.write("  {\n")
            f.write(f"    id: {ts_str(m['id'])},\n    name: {ts_str(m['name'])},\n")
            f.write(f"    age: {m['age']},\n    nationality: {ts_str(m['nationality'])},\n")
            f.write(f"    context: {ts_str(m['context'])},\n    marketPool: {ts_str(m['pool'])},\n")
            f.write(f"    marketStatus: {ts_str(m['status'])},\n    primaryRole: {ts_str(m['role'])},\n")
            f.write(f"    immediateF1Eligible: {ts_bool(m['eligible'])},\n")
            f.write(f"    skills: {skills_ts(m['skills'])},\n")
            f.write(f"    overall: {r1(m['overall'])},\n    potential: {r1(m['potential'])},\n")
            f.write(f"    potentialDelta: {r1(m['potentialDelta'])},\n    developmentRate: {r1(m['devRate'])},\n")
            f.write(f"    f1Readiness: {m['readiness']},\n")
            f.write(f"    salary: {r1(m['salary'])},\n    sponsorValue: {r1(m['sponsor'])},\n    buyoutCost: {r1(m['buyout'])},\n")
            f.write(f"    negotiationDifficulty: {ts_str(m['negDiff'])},\n")
            f.write(f"    suggestedUse: {ts_str(m['use'])},\n    notes: {ts_str(m['notes'])},\n  }},\n")
        f.write("];\n")

    # youth
    with open(ensure("market", f"youthProspects{tag}.ts"), "w") as f:
        f.write(header)
        f.write("import type { YouthProspect } from '../../types/marketTypes';\n\n")
        f.write(f"export const youthProspects{tag}: YouthProspect[] = [\n")
        for p in youth:
            f.write("  {\n")
            f.write(f"    id: {ts_str(p['id'])},\n    name: {ts_str(p['name'])},\n")
            f.write(f"    age: {p['age']},\n    birthYear: {p['birthYear']},\n")
            f.write(f"    nationality: {ts_str(p['nationality'])},\n    currentLevel: {ts_str(p['level'])},\n")
            f.write(f"    marketPool: {ts_str(p['pool'])},\n    marketStatus: {ts_str(p['status'])},\n")
            f.write(f"    academyEligibleNow: {ts_bool(p['eligible'])},\n    earliestFullAcademyYear: {p['earliest']},\n")
            f.write(f"    skills: {skills_ts(p['skills'])},\n")
            f.write(f"    overall: {r1(p['overall'])},\n    potential: {r1(p['potential'])},\n")
            f.write(f"    potentialDelta: {r1(p['potentialDelta'])},\n    developmentRate: {r1(p['devRate'])},\n")
            f.write(f"    yearsUntilF1Ready: {p['years']},\n")
            f.write(f"    signingCost: {r1(p['signing'])},\n    yearlyAcademyCost: {r1(p['academy'])},\n")
            f.write(f"    riskLevel: {ts_str(p['risk'])},\n    suggestedPath: {ts_str(p['path'])},\n    notes: {ts_str(p['notes'])},\n  }},\n")
        f.write("];\n")


def main():
    args = [a.lower() for a in sys.argv[1:]]
    if "indycar" in args:
        gen_indycar(2026)
        return
    years = [int(a) for a in sys.argv[1:] if a.isdigit()]
    if not years and os.environ.get("YEAR"):
        years = [int(os.environ["YEAR"])]
    if not years:
        years = [1997, 1998, 1999, 2000]
    for y in years:
        gen(y)


if __name__ == "__main__":
    main()
