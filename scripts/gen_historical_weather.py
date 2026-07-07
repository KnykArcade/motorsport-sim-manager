#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date as date_cls, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/motorsport-sim-manager')
WORKBOOK = Path('/home/ubuntu/refwb-push/reference-workbooks/master/MASTER_SEASON_TRACK_LIST_WITH_RACE_START_TIMES.xlsx')
CACHE_DIR = ROOT / '.cache' / 'historical-weather'
GENERATED_DIR = ROOT / 'src/data/weather/generated'
PHASE0_DIR = ROOT / 'src/data/phase0/generated'

HEADERS = {
    'User-Agent': 'motorsport-sim-manager-weather-pipeline/1.0',
}


def series_token(series: str) -> str:
    return 'ChampCar' if series == 'Champ Car' else series


def normalize_series(year: int, series: str) -> str:
    if series == 'CART' and 2004 <= year <= 2007:
        return 'Champ Car'
    return series


def default_local_start_time(series: str, year: int) -> str:
    if series == 'F1':
        return '14:00'
    if series == 'Champ Car':
        return '14:00' if year >= 2004 else '13:00'
    return '13:30' if year >= 2008 else '13:00'


def join_key(year: int, series: str, round_no: int) -> str:
    return f'{year}-{series}-{round_no}'


def exact_season_file(year: int, series: str) -> Path:
    return PHASE0_DIR / f'season{year}{series_token(series)}.ts'


def parse_rounds_from_season_file(path: Path) -> set[int]:
    text = path.read_text(encoding='utf-8')
    return {int(v) for v in re.findall(r'"round":\s*(\d+)', text)}


def ts_string(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def build_weather_state(code: int, changing_soon: bool = False) -> dict[str, Any]:
    condition = map_wmo_code(code)
    grip = {
        'Dry': 1,
        'Cloudy': 0.97,
        'Drying': 0.85,
        'Changeable': 0.8,
        'LightRain': 0.72,
        'HeavyRain': 0.5,
    }[condition]
    label = {
        'Dry': 'Dry',
        'Cloudy': 'Cloudy',
        'LightRain': 'Light Rain',
        'HeavyRain': 'Heavy Rain',
        'Drying': 'Drying',
        'Changeable': 'Changeable',
    }[condition]
    return {
        'condition': condition,
        'gripLevel': grip,
        'wet': condition in {'LightRain', 'HeavyRain'},
        'changingSoon': changing_soon,
        'label': label,
    }


def map_wmo_code(code: int) -> str:
    if code == 0:
        return 'Dry'
    if code in {1, 2, 3, 45, 48}:
        return 'Cloudy'
    if code in {51, 53, 55, 56, 57, 61, 63, 80, 81}:
        return 'LightRain'
    if code in {65, 66, 67, 71, 73, 75, 77, 82, 85, 86, 95, 96, 99}:
        return 'HeavyRain'
    return 'Changeable' if 50 <= code < 80 else 'Cloudy'


def offset_local_time(date: str, local_time: str, minutes_offset: int) -> str:
    year, month, day = map(int, date.split('-'))
    hour, minute = map(int, local_time.split(':'))
    # Deliberately use UTC math to avoid environment timezone drift.
    import datetime as dt

    base = dt.datetime(year, month, day, hour, minute, tzinfo=dt.timezone.utc)
    shifted = base + dt.timedelta(minutes=minutes_offset)
    return shifted.strftime('%Y-%m-%dT%H:%M')


def build_timeline(anchor: dict[str, Any], hourly: list[dict[str, Any]], window_minutes: int = 180, resolution_minutes: int = 15) -> list[dict[str, Any]]:
    if not hourly:
        return []
    samples: list[dict[str, Any]] = []
    sample_count = max(1, (window_minutes + resolution_minutes - 1) // resolution_minutes)
    for index in range(sample_count):
        minutes_from_start = index * resolution_minutes
        hour_index = min(len(hourly) - 1, max(0, minutes_from_start // 60))
        current = hourly[hour_index]
        nxt = hourly[min(hour_index + 1, len(hourly) - 1)]
        changing_soon = bool(nxt) and nxt['weatherCode'] != current['weatherCode'] and minutes_from_start % 60 >= 30
        state = build_weather_state(int(current['weatherCode']), changing_soon)
        samples.append({
            'time': offset_local_time(anchor['date'], anchor['localStartTime'], minutes_from_start),
            'weatherCode': int(current['weatherCode']),
            'condition': state['condition'],
            'state': state,
            'precipitationMm': current.get('precipitationMm'),
            'rainMm': current.get('rainMm'),
            'cloudCover': current.get('cloudCover'),
            'temperature2m': current.get('temperature2m'),
            'windSpeed10m': current.get('windSpeed10m'),
        })
    return samples


def fetch_json(url: str, timeout: int = 60) -> dict[str, Any]:
    req = Request(url, headers=HEADERS)
    with urlopen(req, timeout=timeout) as response:
        payload = response.read().decode('utf-8')
    return json.loads(payload)


def fetch_with_backoff(url: str, cache_path: Path) -> dict[str, Any]:
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding='utf-8'))
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    delay = 1.0
    last_error: Exception | None = None
    for attempt in range(6):
        try:
            data = fetch_json(url)
            cache_path.write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')
            return data
        except HTTPError as err:
            last_error = err
            if err.code == 429:
                time.sleep(delay + random.random() * 0.25)
                delay *= 2
                continue
            if 500 <= err.code < 600 and attempt < 5:
                time.sleep(delay + random.random() * 0.25)
                delay *= 2
                continue
            raise
        except URLError as err:
            last_error = err
            time.sleep(delay + random.random() * 0.25)
            delay *= 2
    if last_error:
        raise last_error
    raise RuntimeError('fetch failed without exception')


def fetch_archive(latitude: float, longitude: float, date: str) -> dict[str, Any]:
    params = urlencode({
        'latitude': latitude,
        'longitude': longitude,
        'start_date': date,
        'end_date': date,
        'timezone': 'auto',
        'hourly': 'weather_code,precipitation,rain,cloud_cover,temperature_2m,wind_speed_10m',
        'temperature_unit': 'celsius',
        'wind_speed_unit': 'kmh',
        'precipitation_unit': 'mm',
        'timeformat': 'iso8601',
    })
    return fetch_json(f'https://archive-api.open-meteo.com/v1/archive?{params}')


def parse_hourly_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    hourly = payload.get('hourly') or {}
    times = hourly.get('time') or []
    weather_codes = hourly.get('weather_code') or []
    precipitation = hourly.get('precipitation') or []
    rain = hourly.get('rain') or []
    cloud_cover = hourly.get('cloud_cover') or []
    temperature = hourly.get('temperature_2m') or []
    wind_speed = hourly.get('wind_speed_10m') or []
    points = []
    for idx, time_value in enumerate(times):
        points.append({
            'time': time_value,
            'weatherCode': weather_codes[idx] if idx < len(weather_codes) else 0,
            'precipitationMm': precipitation[idx] if idx < len(precipitation) else None,
            'rainMm': rain[idx] if idx < len(rain) else None,
            'cloudCover': cloud_cover[idx] if idx < len(cloud_cover) else None,
            'temperature2m': temperature[idx] if idx < len(temperature) else None,
            'windSpeed10m': wind_speed[idx] if idx < len(wind_speed) else None,
        })
    return points


@dataclass
class WeatherRow:
    season: int
    source_series: str
    series: str
    round_no: int
    race_name: str
    track_name: str
    track_id: str
    date: str
    latitude: float
    longitude: float
    start_time_local: str
    start_date_for_weather: str
    start_time_confidence: str
    start_time_method: str

    @property
    def race_id(self) -> str:
        return join_key(self.season, self.series, self.round_no)

    @property
    def local_start_time(self) -> str:
        return self.start_time_local or default_local_start_time(self.series, self.season)

    @property
    def weather_date(self) -> str:
        return self.start_date_for_weather or self.date

    @property
    def start_time_source(self) -> str:
        return 'workbook-estimate' if self.start_time_local else 'series-default'

    @property
    def assumptions(self) -> list[str]:
        if self.start_time_local:
            return [
                f'Workbook start time estimate ({self.start_time_confidence or "unknown"}): {self.start_time_local}',
                f'Start time method: {self.start_time_method or "unspecified"}',
            ]
        return [f'No workbook start time; using {self.series} default {self.local_start_time}']


def load_rows() -> list[WeatherRow]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb['Season_Track_Ratings']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: idx for idx, name in enumerate(headers)}
    rows: list[WeatherRow] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append(WeatherRow(
            season=int(values[index['Season']]),
            source_series=str(values[index['Series']]),
            series=normalize_series(int(values[index['Season']]), str(values[index['Series']])),
            round_no=int(values[index['Round']]),
            race_name=str(values[index['Race_Name']]),
            track_name=str(values[index['Track']]),
            track_id=str(values[index['TrackId']]),
            date=str(values[index['Race_Start_Date_For_Weather']] or values[index['Race_Date']]),
            latitude=float(values[index['Weather_Latitude']]),
            longitude=float(values[index['Weather_Longitude']]),
            start_time_local=str(values[index['Race_Start_Time_Local']] or ''),
            start_date_for_weather=str(values[index['Race_Start_Date_For_Weather']] or values[index['Race_Date']]),
            start_time_confidence=str(values[index['Start_Time_Confidence']] or ''),
            start_time_method=str(values[index['Start_Time_Method']] or ''),
        ))
    return rows


def parse_iso_date(value: str) -> date_cls | None:
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def season_export_name(year: int, series: str) -> str:
    return f'season{year}{series_token(series)}Weather'


def write_ts_module(path: Path, header_lines: list[str], export_name: str, type_name: str, data: dict[str, Any]) -> None:
    body = json.dumps(data, ensure_ascii=False, indent=2)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as fh:
        fh.write('// AUTO-GENERATED by scripts/gen_historical_weather.py. Do not hand-edit.\n')
        for line in header_lines:
            fh.write(f'{line}\n')
        fh.write(f'export const {export_name} = {body} satisfies Record<string, {type_name}>;\n')


def load_cached_payload(cache_path: Path) -> dict[str, Any]:
    if not cache_path.exists():
        raise FileNotFoundError(f'Missing cached archive payload: {cache_path}')
    return json.loads(cache_path.read_text(encoding='utf-8'))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--workers', type=int, default=4)
    parser.add_argument('--limit', type=int, default=0)
    args = parser.parse_args()

    rows = load_rows()
    if args.limit and args.limit > 0:
        rows = rows[:args.limit]

    matched_exact = 0
    matched_normalized = 0
    round_cache: dict[tuple[int, str], set[int]] = {}
    for row in rows:
        exact_series = row.source_series
        normalized_key = (row.season, row.series)
        exact_path = exact_season_file(row.season, exact_series)
        normalized_path = exact_season_file(row.season, row.series)
        exact_rounds = round_cache.get((row.season, exact_series))
        if exact_rounds is None and exact_path.exists():
            exact_rounds = parse_rounds_from_season_file(exact_path)
            round_cache[(row.season, exact_series)] = exact_rounds
        normalized_rounds = round_cache.get(normalized_key)
        if normalized_rounds is None and normalized_path.exists():
            normalized_rounds = parse_rounds_from_season_file(normalized_path)
            round_cache[normalized_key] = normalized_rounds
        if exact_rounds and row.round_no in exact_rounds:
            matched_exact += 1
        if normalized_rounds and row.round_no in normalized_rounds:
            matched_normalized += 1

    print(json.dumps({
        'rows': len(rows),
        'exactJoinMatches': matched_exact,
        'normalizedJoinMatches': matched_normalized,
        'exactJoinMatchRate': round(matched_exact / len(rows), 6) if rows else 0,
        'normalizedJoinMatchRate': round(matched_normalized / len(rows), 6) if rows else 0,
    }, indent=2))

    # Note: the workbook stores CART for the Champ Car era; we normalize for the game runtime.
    fetched_rows = []
    skipped_rows: list[dict[str, Any]] = []
    start_time_counts = {'real': 0, 'fallback': 0}
    confidence_counts: dict[str, int] = defaultdict(int)
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        futures = {executor.submit(_process_row, row): row for row in rows}
        for future in as_completed(futures):
            result = future.result()
            if result.get('skipped'):
                skipped_rows.append(result['skip'])
            fetched_rows.append(result)
            meta = result['meta']
            if meta.get('startTimeSource') == 'series-default':
                start_time_counts['fallback'] += 1
            else:
                start_time_counts['real'] += 1
                confidence = str(meta.get('startTimeConfidence') or 'unknown')
                confidence_counts[confidence] += 1

    fetched_rows.sort(key=lambda item: (item['meta']['seasonYear'], item['meta']['series'], item['meta']['round']))

    race_meta: dict[str, Any] = {}
    track_coords: dict[str, Any] = {}
    season_groups: dict[tuple[int, str], dict[str, Any]] = defaultdict(dict)
    season_counts: dict[tuple[int, str], int] = defaultdict(int)
    timeline_counts = 0
    for item in fetched_rows:
        meta = item['meta']
        timeline = item['timeline']
        race_meta[meta['raceId']] = meta
        track_coords.setdefault(meta['trackId'], {
            'trackId': meta['trackId'],
            'latitude': meta['latitude'],
            'longitude': meta['longitude'],
            'sourceRaceId': meta['raceId'],
        })
        season_groups[(meta['seasonYear'], meta['series'])][meta['raceId']] = timeline
        season_counts[(meta['seasonYear'], meta['series'])] += 1
        timeline_counts += 1

    write_ts_module(
        GENERATED_DIR / 'raceMeta.ts',
        ["import type { HistoricalWeatherRaceMeta } from '../weatherTypes';"],
        'historicalWeatherRaceMeta',
        'HistoricalWeatherRaceMeta',
        race_meta,
    )
    write_ts_module(
        GENERATED_DIR / 'trackCoordinates.ts',
        ["import type { HistoricalWeatherTrackCoordinate } from '../weatherTypes';"],
        'historicalWeatherTrackCoordinates',
        'HistoricalWeatherTrackCoordinate',
        track_coords,
    )

    total_bytes = 0
    for (season_year, series), timelines in sorted(season_groups.items()):
        export_name = season_export_name(season_year, series)
        path = GENERATED_DIR / f'season{season_year}{series_token(series)}.ts'
        write_ts_module(
            path,
            ["import type { HistoricalWeatherTimeline } from '../../../sim/historicalWeather';"],
            export_name,
            'HistoricalWeatherTimeline',
            timelines,
        )
        total_bytes += path.stat().st_size

    manifest = {
        'rows': len(rows),
        'startTimeRealCount': start_time_counts['real'],
        'startTimeFallbackCount': start_time_counts['fallback'],
        'startTimeConfidenceCounts': dict(sorted(confidence_counts.items())),
        'skippedCount': len(skipped_rows),
        'skippedRows': skipped_rows,
        'raceMetaCount': len(race_meta),
        'trackCoordinateCount': len(track_coords),
        'timelineCount': timeline_counts,
        'seasonCount': len(season_groups),
        'seasonCounts': {f'{year}-{series}': count for (year, series), count in sorted(season_counts.items())},
        'exactJoinMatches': matched_exact,
        'normalizedJoinMatches': matched_normalized,
        'rawCacheDir': str(CACHE_DIR),
        'generatedBytes': total_bytes,
    }
    (GENERATED_DIR / 'manifest.json').write_text(json.dumps(manifest, indent=2), encoding='utf-8')
    print(json.dumps(manifest, indent=2))
    return 0


def _process_row(row: WeatherRow) -> dict[str, Any]:
    parsed_date = parse_iso_date(row.weather_date)
    base_assumptions = list(row.assumptions)
    if row.weather_date != row.date:
        base_assumptions.append(f'Weather date for archive: {row.weather_date} (Race_Start_Date_For_Weather)')
    if parsed_date is None:
        reason = f'Workbook weather date "{row.weather_date}" is not ISO-8601; archive fetch skipped'
        meta = {
            'raceId': row.race_id,
            'seasonYear': row.season,
            'series': row.series,
            'sourceSeries': row.source_series,
            'round': row.round_no,
            'trackId': row.track_id,
            'trackName': row.track_name,
            'date': row.weather_date,
            'localStartTime': row.local_start_time,
            'timezone': 'auto',
            'latitude': row.latitude,
            'longitude': row.longitude,
            'coordinateSource': 'workbook',
            'startTimeSource': row.start_time_source,
            'startTimeConfidence': row.start_time_confidence or None,
            'startTimeMethod': row.start_time_method or None,
            'assumptions': base_assumptions + [reason],
        }
        return {
            'skipped': True,
            'skip': {
                'raceId': row.race_id,
                'reason': reason,
                'date': row.weather_date,
            },
            'meta': meta,
            'timeline': {
                'anchor': {
                    'raceId': row.race_id,
                    'raceName': row.race_name,
                    'trackId': row.track_id,
                    'trackName': row.track_name,
                    'year': row.season,
                    'series': row.series,
                    'date': row.weather_date,
                    'localStartTime': row.local_start_time,
                    'timezone': 'auto',
                    'latitude': row.latitude,
                    'longitude': row.longitude,
                    'coordinateSource': 'workbook',
                    'startTimeSource': row.start_time_source,
                    'startTimeConfidence': row.start_time_confidence or None,
                    'startTimeMethod': row.start_time_method or None,
                },
                'source': 'open-meteo-archive',
                'resolutionMinutes': 15,
                'assumptions': base_assumptions + [reason],
                'samples': [],
            },
        }
    cutoff = date_cls.today()
    if parsed_date > cutoff:
        reason = f'Workbook weather date {row.weather_date} is after archive cutoff {cutoff.isoformat()}'
        meta = {
            'raceId': row.race_id,
            'seasonYear': row.season,
            'series': row.series,
            'sourceSeries': row.source_series,
            'round': row.round_no,
            'trackId': row.track_id,
            'trackName': row.track_name,
            'date': row.weather_date,
            'localStartTime': row.local_start_time,
            'timezone': 'auto',
            'latitude': row.latitude,
            'longitude': row.longitude,
            'coordinateSource': 'workbook',
            'startTimeSource': row.start_time_source,
            'startTimeConfidence': row.start_time_confidence or None,
            'startTimeMethod': row.start_time_method or None,
            'assumptions': base_assumptions + [reason],
        }
        return {
            'skipped': True,
            'skip': {
                'raceId': row.race_id,
                'reason': reason,
                'date': row.weather_date,
            },
            'meta': meta,
            'timeline': {
                'anchor': {
                    'raceId': row.race_id,
                    'raceName': row.race_name,
                    'trackId': row.track_id,
                    'trackName': row.track_name,
                    'year': row.season,
                    'series': row.series,
                    'date': row.weather_date,
                    'localStartTime': row.local_start_time,
                    'timezone': 'auto',
                    'latitude': row.latitude,
                    'longitude': row.longitude,
                    'coordinateSource': 'workbook',
                    'startTimeSource': row.start_time_source,
                    'startTimeConfidence': row.start_time_confidence or None,
                    'startTimeMethod': row.start_time_method or None,
                },
                'source': 'open-meteo-archive',
                'resolutionMinutes': 15,
                'assumptions': base_assumptions + [reason],
                'samples': [],
            },
        }
    cache_path = CACHE_DIR / f'{row.race_id}.json'
    payload = load_cached_payload(cache_path)
    hourly = parse_hourly_payload(payload)
    meta = {
        'raceId': row.race_id,
        'seasonYear': row.season,
        'series': row.series,
        'sourceSeries': row.source_series,
        'round': row.round_no,
        'trackId': row.track_id,
        'trackName': row.track_name,
        'date': row.weather_date,
        'localStartTime': row.local_start_time,
        'timezone': 'auto',
        'latitude': row.latitude,
        'longitude': row.longitude,
        'coordinateSource': 'workbook',
        'startTimeSource': row.start_time_source,
        'startTimeConfidence': row.start_time_confidence or None,
        'startTimeMethod': row.start_time_method or None,
        'assumptions': base_assumptions,
    }
    timeline = {
        'anchor': {
            'raceId': row.race_id,
            'raceName': row.race_name,
            'trackId': row.track_id,
            'trackName': row.track_name,
            'year': row.season,
            'series': row.series,
            'date': row.weather_date,
            'localStartTime': row.local_start_time,
            'timezone': 'auto',
            'latitude': row.latitude,
            'longitude': row.longitude,
            'coordinateSource': 'workbook',
            'startTimeSource': row.start_time_source,
            'startTimeConfidence': row.start_time_confidence or None,
            'startTimeMethod': row.start_time_method or None,
        },
        'source': 'open-meteo-archive',
        'resolutionMinutes': 15,
        'assumptions': base_assumptions,
        'samples': build_timeline({
            'date': row.weather_date,
            'localStartTime': row.local_start_time,
        }, hourly, 180, 15),
    }
    return {'meta': meta, 'timeline': timeline}


def fetch_archive_url(latitude: float, longitude: float, date: str) -> str:
    params = urlencode({
        'latitude': latitude,
        'longitude': longitude,
        'start_date': date,
        'end_date': date,
        'timezone': 'auto',
        'hourly': 'weather_code,precipitation,rain,cloud_cover,temperature_2m,wind_speed_10m',
        'temperature_unit': 'celsius',
        'wind_speed_unit': 'kmh',
        'precipitation_unit': 'mm',
        'timeformat': 'iso8601',
    })
    return f'https://archive-api.open-meteo.com/v1/archive?{params}'


if __name__ == '__main__':
    raise SystemExit(main())
