import openpyxl, sys

path = sys.argv[1] if len(sys.argv) > 1 else "C:/Users/tnick/OneDrive/Desktop/F1 Rating/F1/F1_2016_Season_Full_Update.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

# Find calendar sheet
for name in wb.sheetnames:
    if "calendar" in name.lower() or "track" in name.lower():
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        # Print ALL columns from header row
        header = rows[0]
        print(f"=== {name} - ALL columns ({len(header)} total) ===")
        for i, c in enumerate(header):
            print(f"  Col {i}: {c}")
        # Print all data rows fully
        print(f"\n=== All data rows ===")
        for i, row in enumerate(rows[1:], 1):
            vals = [str(c)[:40] if c else '' for c in row]
            print(f"  Row {i}: {vals}")
        break
