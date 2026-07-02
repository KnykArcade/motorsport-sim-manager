"""Fix duplicate track objects in track TS files by removing duplicate entries.
Also fix IndyCar 2008-2015 calendar laps with historical data.
"""
import re
import os

ROOT = os.path.join(os.path.dirname(__file__), '..', 'src', 'data')
MI_KM = 1.60934

# IndyCar historical lap counts for 2008-2015
# Format: year -> { round: laps }
INDYCAR_LAPS = {
    2008: {1:80, 2:200, 3:228, 4:100, 5:83, 6:200, 7:225, 8:228, 9:250, 10:300, 11:60, 12:200, 13:85, 14:65, 15:200, 16:75, 17:90, 18:200},
    2009: {1:100, 2:200, 3:223, 4:100, 5:83, 6:200, 7:225, 8:228, 9:250, 10:300, 11:60, 12:200, 13:85, 14:95, 15:200, 16:75, 17:200},
    2010: {1:100, 2:200, 3:200, 4:100, 5:83, 6:200, 7:200, 8:228, 9:250, 10:300, 11:60, 12:200, 13:85, 14:95, 15:200, 16:75, 17:200},
    2011: {1:100, 2:200, 3:100, 4:100, 5:83, 6:200, 7:55, 8:114, 9:250, 10:300, 11:85, 12:50, 13:85, 14:90, 15:200, 16:75, 17:200},
    2012: {1:100, 2:200, 3:118, 4:90, 5:85, 6:200, 7:60, 8:250, 9:250, 10:15, 11:80, 12:55, 13:85, 14:60, 15:200},
    2013: {1:100, 2:200, 3:118, 4:90, 5:85, 6:200, 7:70, 8:55, 9:250, 10:250, 11:15, 12:85, 13:90, 14:55, 15:85, 16:50, 17:200, 18:75, 19:200},
    2014: {1:100, 2:200, 3:82, 4:90, 5:85, 6:200, 7:70, 8:55, 9:250, 10:250, 11:300, 12:85, 13:90, 14:55, 15:85, 16:60, 17:200, 18:55},
    2015: {1:100, 2:200, 3:50, 4:90, 5:85, 6:200, 7:70, 8:55, 9:250, 10:300, 11:85, 12:90, 13:55, 14:85, 15:200, 16:55},
}


def fix_indycar_2008_2015_laps():
    """Add lap counts to IndyCar 2008-2015 calendars using historical data."""
    for year in range(2008, 2016):
        season_file = os.path.join(ROOT, "seasons", f"season{year}IndyCar.ts")
        if not os.path.exists(season_file):
            continue
        content = open(season_file, encoding="utf-8").read()
        laps_data = INDYCAR_LAPS.get(year, {})

        # Also read track lengths from the tracks file
        tracks_file = os.path.join(ROOT, "tracks", f"tracks{year}IndyCar.ts")
        track_lengths = {}
        if os.path.exists(tracks_file):
            tc = open(tracks_file, encoding="utf-8").read()
            # Try to extract track length from the track file
            # Track files don't have km field directly, but we can read from Excel
            pass

        # Read track lengths from Excel
        import openpyxl
        excel_path = f"C:/Users/tnick/OneDrive/Desktop/F1 Rating/IndyCar/IndyCar_{year}_Season_Full_Update.xlsx"
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            for name in wb.sheetnames:
                if "calendar" in name.lower():
                    ws = wb[name]
                    rows = list(ws.iter_rows(values_only=True))
                    header = rows[0]
                    col_idx = {}
                    for i, c in enumerate(header):
                        if c and str(c).strip():
                            col_idx[str(c).strip().lower()] = i

                    for row in rows[1:]:
                        if not row or all(c is None or str(c).strip() == "" for c in row):
                            continue
                        rnd = row[col_idx.get("round", 0)] if "round" in col_idx else None
                        if rnd is None:
                            continue
                        rnd = int(float(rnd))
                        miles = None
                        for mkey in ["track length mi", "track length miles", "length mi", "miles"]:
                            if mkey in col_idx:
                                mi = row[col_idx[mkey]]
                                if mi:
                                    try:
                                        miles = float(mi)
                                    except:
                                        pass
                                break
                        if miles:
                            track_lengths[rnd] = round(miles * MI_KM, 1)
                    break
            wb.close()

        # Fix each race entry
        def fix_race(match):
            full = match.group(0)
            round_match = re.search(r"round: (\d+)", full)
            if not round_match:
                return full
            rnd = int(round_match.group(1))
            laps = laps_data.get(rnd, 0)
            track_km = track_lengths.get(rnd, 0)
            dist_km = round(laps * track_km, 1) if laps and track_km else None

            full = re.sub(r"laps: \d+", f"laps: {laps}", full)
            if dist_km is not None:
                full = re.sub(r"distanceKm: (undefined|\d+\.?\d*)", f"distanceKm: {dist_km}", full)
            return full

        content = re.sub(r"\{[^}]*id: 'r-\d+-\d+'[^}]*\}", fix_race, content)
        open(season_file, "w", encoding="utf-8").write(content)
        print(f"Fixed IndyCar {year} laps: {len(laps_data)} races")


def remove_duplicate_tracks():
    """Remove duplicate track objects from track TS files."""
    from collections import Counter

    for year in range(1990, 2027):
        for suffix in ["", "IndyCar"]:
            f = os.path.join(ROOT, "tracks", f"tracks{year}{suffix}.ts")
            if not os.path.exists(f):
                continue

            content = open(f, encoding="utf-8").read()

            # Find all track block boundaries
            # Each track block starts with "  {\n" and ends with "  },\n" or "  }\n"
            # We need to handle nested braces in setupProfile

            # Find all top-level track objects
            track_pattern = re.compile(r"^  \{$", re.MULTILINE)
            starts = [m.start() for m in track_pattern.finditer(content)]

            if not starts:
                continue

            # Find matching closing brace for each track block
            blocks = []
            for start in starts:
                depth = 1
                pos = start + 4  # skip "  {\n"
                while pos < len(content) and depth > 0:
                    if content[pos] == '{':
                        depth += 1
                    elif content[pos] == '}':
                        depth -= 1
                    pos += 1
                # Find the end of this block (including trailing comma and newline)
                end = pos
                while end < len(content) and content[end] in ',\n\r':
                    end += 1
                block_text = content[start:end]
                # Extract track ID
                id_match = re.search(r"id: '([^']+)'", block_text)
                if id_match:
                    blocks.append((start, end, id_match.group(1), block_text))

            # Find duplicates
            ids = [b[2] for b in blocks]
            counts = Counter(ids)
            dups = {k: v for k, v in counts.items() if v > 1}
            if not dups:
                continue

            # Keep only first occurrence of each ID
            seen = set()
            remove_ranges = []
            for start, end, tid, block_text in blocks:
                if tid in seen:
                    remove_ranges.append((start, end))
                else:
                    seen.add(tid)

            if not remove_ranges:
                continue

            # Remove duplicate blocks (work backwards to preserve indices)
            new_content = content
            for start, end in sorted(remove_ranges, reverse=True):
                new_content = new_content[:start] + new_content[end:]

            open(f, "w", encoding="utf-8").write(new_content)
            print(f"Fixed tracks{year}{suffix}: removed {len(remove_ranges)} duplicate track objects ({list(dups.keys())})")


if __name__ == "__main__":
    print("=== Fixing IndyCar 2008-2015 laps ===")
    fix_indycar_2008_2015_laps()

    print("\n=== Removing duplicate track objects ===")
    remove_duplicate_tracks()

    print("\nDone")
