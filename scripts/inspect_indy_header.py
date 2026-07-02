import openpyxl

path = "C:/Users/tnick/OneDrive/Desktop/F1 Rating/IndyCar/IndyCar_2016_Season_Full_Update.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

for name in wb.sheetnames:
    if "calendar" in name.lower() or "track" in name.lower():
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        print(f"=== {name} - ALL columns ({len(header)} total) ===")
        for i, c in enumerate(header):
            print(f"  Col {i}: '{c}'")
        break
