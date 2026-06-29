"""Generate clean TypeScript seed data from the 1995 workbook.

Run once to (re)produce the data/*1995*.ts files. The Excel file is NOT parsed
at runtime; this is a build-time conversion step kept for reproducibility.
"""
import os
import re
import unicodedata

import openpyxl

XLSX = os.environ.get(
    "XLSX",
    "/home/ubuntu/attachments/432621e8-eb6e-4123-95c5-10853d30f6b1/F1_1995_Track_Driver_Car_Performance.xlsx",
)
OUT = os.environ.get("OUT", "/home/ubuntu/repos/motorsport-sim-manager/src/data")


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace("\xa0", "").strip()
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return None


def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s


def r1(x):
    return round(x * 10) / 10


def ts_str(s):
    if s is None:
        s = ""
    s = str(s).replace("\xa0", " ").strip()
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


wb = openpyxl.load_workbook(XLSX, data_only=True)
ratings = wb["Ratings"]
carperf = wb["CarPerformance"]

rows = list(ratings.iter_rows(values_only=True))
header = rows[0]

# ---- Tracks (rows until a blank line) ----
tracks = []
for row in rows[1:]:
    if row[0] is None:
        break
    (number, gp, track, km, laps, corners, braking, straights, traction,
     elev, technical, overtaking, surface, risk, endurance, primary,
     downforce, topspeed, mech, brake, relrisk, strategy, notes,
     aero_d, power_d, mech_d, risk_d, archetype) = row[:28]
    tid = slug(track)
    tracks.append({
        "id": tid,
        "name": track,
        "gpName": gp,
        "round": number,
        "km": num(km),
        "laps": num(laps),
        "attrs": {
            "corners": num(corners), "braking": num(braking), "straights": num(straights),
            "tractionAcceleration": num(traction), "elevationBlindCorners": num(elev),
            "technical": num(technical), "overtakingRacecraft": num(overtaking),
            "surfaceGripBumpiness": num(surface), "riskWallProximity": num(risk),
            "enduranceConsistency": num(endurance),
        },
        "setup": {
            "primarySetupProfile": primary, "downforceLevel": downforce,
            "topSpeedEmphasis": topspeed, "mechanicalGripEmphasis": mech,
            "brakeDemand": brake, "reliabilityRiskFocus": relrisk,
            "strategyNotes": strategy,
            "aeroDemand": num(aero_d), "powerDemand": num(power_d),
            "mechanicalDemand": num(mech_d), "riskDemand": num(risk_d),
        },
        "archetype": archetype,
        "notes": notes,
    })

# topSpeedEmphasis/mech/brake/relrisk are word scales -> map to 1-10
WORD = {
    "very low": 1, "minimum": 1, "minimum-low": 2, "low": 3, "low-medium": 4,
    "medium-low": 4, "medium": 5, "medium-high": 6, "high": 7, "high-maximum": 8,
    "maximum": 9, "very high": 10,
}


def wordnum(v):
    if v is None:
        return 5
    return WORD.get(str(v).strip().lower(), 5)


# ---- Drivers (after the blank rows, a second header) ----
driver_rows = []
seen_header = False
for row in rows:
    if row and row[0] == "Team":
        seen_header = True
        continue
    if seen_header and row and row[0]:
        driver_rows.append(row)

# ---- Car performance ----
car_rows = []
for row in carperf.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    car_rows.append(row)

# Display-name + color mapping for teams (data uses some informal spellings).
TEAM_META = {
    "Benetton": ("Benetton", "BEN", "United Kingdom", "#1Fa05a"),
    "Williams": ("Williams", "WIL", "United Kingdom", "#1f6fd6"),
    "Ferrari": ("Ferrari", "FER", "Italy", "#d4001a"),
    "Mclaren": ("McLaren", "MCL", "United Kingdom", "#ff8000"),
    "Liger": ("Ligier", "LIG", "France", "#2f8fe0"),
    "Sauber": ("Sauber", "SAU", "Switzerland", "#00a3d6"),
    "Tyrell": ("Tyrrell", "TYR", "United Kingdom", "#1f5fd0"),
    "Jordan": ("Jordan", "JOR", "Ireland", "#e6b800"),
    "Minardi": ("Minardi", "MIN", "Italy", "#7a7f8a"),
    "Forti": ("Forti", "FOR", "Italy", "#d6c200"),
    "Pacific": ("Pacific", "PAC", "United Kingdom", "#2aa198"),
    "Footwork": ("Footwork", "FTW", "United Kingdom", "#cc2a5e"),
}

# Parse driver number + name e.g. "#1 M. Schumacher"
drivers = []
for row in driver_rows:
    team_raw = row[0]
    raw_name = str(row[1]).strip()
    m = re.match(r"#?\s*(\d+)\s+(.*)", raw_name)
    number = int(m.group(1)) if m else 0
    name = m.group(2).strip() if m else raw_name
    (corners, braking, straights, traction, elev, technical, overtaking,
     surface, risk, endurance, overall) = [num(x) for x in row[2:13]]
    drivers.append({
        "team_raw": team_raw, "number": number, "name": name,
        "corners": corners, "braking": braking, "straights": straights,
        "traction": traction, "elev": elev, "technical": technical,
        "overtaking": overtaking, "surface": surface, "risk": risk,
        "endurance": endurance, "overall": overall,
    })

cars = []
for row in car_rows:
    team_raw, eng, aero, mech, rel, pit = row[:6]
    cars.append({
        "team_raw": team_raw, "enginePower": num(eng), "aeroEfficiency": num(aero),
        "mechanicalGrip": num(mech), "reliability": num(rel), "pitCrewOperations": num(pit),
    })

car_by_team = {c["team_raw"]: c for c in cars}

# Rank teams by car overall for expectedStanding / budget tiers.
team_score = {c["team_raw"]: c["enginePower"] + c["aeroEfficiency"] + c["mechanicalGrip"] + c["reliability"] + c["pitCrewOperations"] for c in cars}
ranked = sorted(team_score, key=lambda t: -team_score[t])
expected = {t: i + 1 for i, t in enumerate(ranked)}

# Budgets: top teams richer.
def budget_for(rank):
    return int(round((90 - (rank - 1) * 6) )) * 1_000_000  # 90M down to ~24M


def did(d):
    return f"d-{d['number']}-{slug(d['name'])}"


def tid(team_raw):
    return "t-" + slug(TEAM_META[team_raw][0])


# ------------------------------------------------------------------ emit TS
os.makedirs(OUT, exist_ok=True)
HEADER = "// AUTO-GENERATED from F1_1995_Track_Driver_Car_Performance.xlsx by scripts/gen_seed.py.\n// Edit ratings here directly; this file is the runtime source of truth.\n\n"

# tracks
with open(os.path.join(OUT, "tracks", "tracks1995.ts"), "w") as f:
    f.write(HEADER)
    f.write("import type { Track } from '../../types/gameTypes';\n\n")
    f.write("export const tracks1995: Track[] = [\n")
    for t in tracks:
        a = t["attrs"]
        s = t["setup"]
        f.write("  {\n")
        f.write(f"    id: {ts_str(t['id'])},\n")
        f.write(f"    name: {ts_str(t['name'])},\n")
        f.write(f"    gpName: {ts_str(t['gpName'])},\n")
        f.write(f"    archetype: {ts_str(t['archetype'])},\n")
        f.write("    attributes: {\n")
        for k, v in a.items():
            f.write(f"      {k}: {v},\n")
        f.write("    },\n")
        f.write("    setupProfile: {\n")
        f.write(f"      primarySetupProfile: {ts_str(s['primarySetupProfile'])},\n")
        f.write(f"      downforceLevel: {ts_str(s['downforceLevel'])},\n")
        f.write(f"      topSpeedEmphasis: {wordnum(s['topSpeedEmphasis'])},\n")
        f.write(f"      mechanicalGripEmphasis: {wordnum(s['mechanicalGripEmphasis'])},\n")
        f.write(f"      brakeDemand: {wordnum(s['brakeDemand'])},\n")
        f.write(f"      reliabilityRiskFocus: {wordnum(s['reliabilityRiskFocus'])},\n")
        f.write(f"      strategyNotes: {ts_str(s['strategyNotes'])},\n")
        f.write(f"      aeroDemand: {s['aeroDemand']},\n")
        f.write(f"      powerDemand: {s['powerDemand']},\n")
        f.write(f"      mechanicalDemand: {s['mechanicalDemand']},\n")
        f.write(f"      riskDemand: {s['riskDemand']},\n")
        f.write("    },\n")
        f.write(f"    ratingNotes: {ts_str(t['notes'])},\n")
        f.write("  },\n")
    f.write("];\n")

# season / calendar
with open(os.path.join(OUT, "seasons", "season1995.ts"), "w") as f:
    f.write(HEADER)
    f.write("import type { Race, Season } from '../../types/gameTypes';\n")
    f.write("import { tracks1995 } from '../tracks/tracks1995';\n\n")
    f.write("export const calendar1995: Race[] = [\n")
    for t in sorted(tracks, key=lambda x: x["round"]):
        dist = r1(t["km"] * t["laps"]) if (t["km"] and t["laps"]) else "undefined"
        f.write("  {\n")
        f.write(f"    id: {ts_str('r-1995-' + str(t['round']))},\n")
        f.write(f"    round: {t['round']},\n")
        f.write(f"    gpName: {ts_str(t['gpName'])},\n")
        f.write(f"    trackId: {ts_str(t['id'])},\n")
        f.write(f"    trackName: {ts_str(t['name'])},\n")
        f.write(f"    laps: {t['laps']},\n")
        f.write(f"    distanceKm: {dist},\n")
        f.write("    completed: false,\n")
        f.write("  },\n")
    f.write("];\n\n")
    f.write("void tracks1995;\n\n")
    f.write("export const season1995: Season = {\n")
    f.write("  id: 's-1995-f1',\n  year: 1995,\n  name: '1995 Formula 1 World Championship',\n")
    f.write("  series: 'F1',\n  calendar: calendar1995,\n  pointsSystemId: 'pts-1995',\n")
    f.write("  regulationSetId: 'reg-1995',\n};\n")

# teams
with open(os.path.join(OUT, "teams", "teams1995.ts"), "w") as f:
    f.write(HEADER)
    f.write("import type { Team } from '../../types/gameTypes';\n\n")
    f.write("export const teams1995: Team[] = [\n")
    diff_by_rank = lambda r: 'Easy' if r <= 3 else ('Medium' if r <= 6 else ('Hard' if r <= 9 else 'Very Hard'))
    for team_raw in ranked:
        name, short, country, color = TEAM_META[team_raw]
        rank = expected[team_raw]
        team_drivers = [d for d in drivers if d["team_raw"] == team_raw]
        team_drivers.sort(key=lambda d: d["number"])
        driver_ids = ", ".join(ts_str(did(d)) for d in team_drivers)
        rep = max(20, 100 - (rank - 1) * 7)
        f.write("  {\n")
        f.write(f"    id: {ts_str(tid(team_raw))},\n")
        f.write(f"    name: {ts_str(name)},\n")
        f.write(f"    shortName: {ts_str(short)},\n")
        f.write(f"    country: {ts_str(country)},\n")
        f.write(f"    carId: {ts_str('car-' + slug(name) + '-1995')},\n")
        f.write(f"    driverIds: [{driver_ids}],\n")
        f.write(f"    budget: {budget_for(rank)},\n")
        f.write(f"    reputation: {rep},\n")
        f.write("    morale: 65,\n")
        f.write(f"    expectedStanding: {rank},\n")
        f.write(f"    difficulty: {ts_str(diff_by_rank(rank))},\n")
        f.write(f"    color: {ts_str(color)},\n")
        f.write("  },\n")
    f.write("];\n")

# cars
with open(os.path.join(OUT, "cars", "cars1995.ts"), "w") as f:
    f.write(HEADER)
    f.write("import type { Car } from '../../types/gameTypes';\n\n")
    f.write("export const cars1995: Car[] = [\n")
    for team_raw in ranked:
        name = TEAM_META[team_raw][0]
        c = car_by_team[team_raw]
        f.write("  {\n")
        f.write(f"    id: {ts_str('car-' + slug(name) + '-1995')},\n")
        f.write(f"    teamId: {ts_str(tid(team_raw))},\n")
        f.write("    seasonYear: 1995,\n")
        f.write("    ratings: {\n")
        for k in ["enginePower", "aeroEfficiency", "mechanicalGrip", "reliability", "pitCrewOperations"]:
            f.write(f"      {k}: {c[k]},\n")
        f.write("    },\n")
        f.write("    condition: 100,\n")
        f.write("    developmentLevel: { enginePower: 0, aeroEfficiency: 0, mechanicalGrip: 0, reliability: 0, pitCrewOperations: 0 },\n")
        f.write("  },\n")
    f.write("];\n")

# drivers
with open(os.path.join(OUT, "drivers", "drivers1995.ts"), "w") as f:
    f.write(HEADER)
    f.write("import type { Driver } from '../../types/gameTypes';\n\n")
    f.write("export const drivers1995: Driver[] = [\n")
    for d in sorted(drivers, key=lambda x: -x["overall"]):
        ov = d["overall"]
        qualifying = r1((ov + d["corners"] + d["braking"]) / 3)
        racePace = r1((ov + d["endurance"] + d["traction"]) / 3)
        adaptability = r1((ov + d["technical"] + d["surface"]) / 3)
        aggression = r1((d["overtaking"] + (11 - d["risk"])) / 2)
        composure = r1((d["risk"] + d["endurance"] + ov) / 3)
        f.write("  {\n")
        f.write(f"    id: {ts_str(did(d))},\n")
        f.write(f"    name: {ts_str(d['name'])},\n")
        f.write(f"    number: {d['number']},\n")
        f.write(f"    teamId: {ts_str(tid(d['team_raw']))},\n")
        f.write("    ratings: {\n")
        f.write(f"      cornering: {d['corners']},\n")
        f.write(f"      braking: {d['braking']},\n")
        f.write(f"      straights: {d['straights']},\n")
        f.write(f"      tractionAcceleration: {d['traction']},\n")
        f.write(f"      elevationBlindCorners: {d['elev']},\n")
        f.write(f"      technical: {d['technical']},\n")
        f.write(f"      overtakingRacecraft: {d['overtaking']},\n")
        f.write(f"      surfaceGripBumpiness: {d['surface']},\n")
        f.write(f"      riskManagement: {d['risk']},\n")
        f.write(f"      enduranceConsistency: {d['endurance']},\n")
        f.write(f"      qualifying: {qualifying},\n")
        f.write(f"      racePace: {racePace},\n")
        f.write(f"      adaptability: {adaptability},\n")
        f.write(f"      aggression: {aggression},\n")
        f.write(f"      composure: {composure},\n")
        f.write(f"      overall: {ov},\n")
        f.write("    },\n")
        f.write("    morale: 65,\n    confidence: 65,\n    traits: [],\n")
        f.write("  },\n")
    f.write("];\n")

print("Generated tracks/season/teams/cars/drivers for 1995.")
print("teams:", len(ranked), "drivers:", len(drivers), "tracks:", len(tracks))
