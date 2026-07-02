import openpyxl

for year in [2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015]:
    path = f"C:/Users/tnick/OneDrive/Desktop/F1 Rating/IndyCar/IndyCar_{year}_Season_Full_Update.xlsx"
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except:
        print(f"{year}: FILE NOT FOUND")
        continue
    for name in wb.sheetnames:
        if "calendar" in name.lower() or "track" in name.lower():
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            print(f"\n{year} ({name}) columns ({len(header)}):")
            for i, c in enumerate(header):
                print(f"  Col {i}: '{c}'")
            break
    wb.close()
