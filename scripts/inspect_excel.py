import openpyxl, sys

path = sys.argv[1] if len(sys.argv) > 1 else "C:/Users/tnick/OneDrive/Desktop/F1 Rating/F1/F1_2016_Season_Full_Update.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

print(f"=== Sheets in {path} ===")
for name in wb.sheetnames:
    print(f"  {name}")

# Find calendar sheet
for name in wb.sheetnames:
    if "calendar" in name.lower() or "track" in name.lower():
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n=== Sheet: {name} ({len(rows)} rows) ===")
        # Find header row
        for i, row in enumerate(rows[:5]):
            print(f"  Row {i}: {[str(c)[:30] if c else '' for c in row[:15]]}")
        # Print first data row
        for i, row in enumerate(rows[1:4], 1):
            print(f"  Data {i}: {[str(c)[:30] if c else '' for c in row[:15]]}")
        break
